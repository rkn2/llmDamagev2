"""One-off script to fill the damage-assessment block (cols 100-154, damage_status,
Notes) for 5 State St rows in Montpelier_Flood_DataInput.xlsx, using
address_assessments.json results plus direct review of the before/after photos
for each address (no external/web data sources).
"""
import openpyxl

XLSX = "Montpelier_Flood_DataInput.xlsx"

wb = openpyxl.load_workbook(XLSX)
ws = wb["BuidlingAttributes"]
headers = [c.value for c in ws[1]]
col = {h: i for i, h in enumerate(headers)}


def cardinal_visibility(front_orientation, vis):
    """Map front/back/left/right photo visibility to N/S/E/W given front_elevation_orientation."""
    if front_orientation == "n":
        return {"n": vis["front"], "s": vis["back"], "e": vis["left"], "w": vis["right"]}
    elif front_orientation == "s":
        return {"n": vis["back"], "s": vis["front"], "e": vis["right"], "w": vis["left"]}
    raise ValueError(front_orientation)


# row -> (flood_height_ft, visibility dict, notes)
# front_elevation_orientation is read from the existing spreadsheet column (not looked up).
ROWS = {
    30: (2.5, {"front": False, "back": False, "left": False, "right": False},
         "damage_status/flood_height from address_assessments.json pipeline (DS2, ~2.5ft "
         "exterior water depth, medium confidence). All available after-photos "
         "(2023-07-11, 2023-07-13, 2024-06-12) are aerial/oblique with no street-level "
         "facade photos for any elevation - wall/cladding/fenestration columns marked "
         "'un' for all sides. The mansard roof appears structurally intact and unchanged "
         "between the 2023 during-flood aerial and the 2024 aerial. Foundation/basement "
         "not visible in any available photo."),
    56: (2.5, {"front": True, "back": False, "left": False, "right": False},
         "damage_status/flood_height from address_assessments.json pipeline (DS2, ~2.5ft "
         "exterior water depth, medium confidence). A 2023-07-11 aerial shows the "
         "building surrounded by floodwater at street level during the event. A "
         "2024-06-12 street-level front (N) photo shows the brick facade and mansard "
         "roof intact with no visible structural, cladding, or fenestration damage; "
         "back/left/right not covered by after-photos (the only other after-photo, "
         "2023-11-1, is too distant/blurry to assess any facade). Foundation/basement "
         "not visually inspected."),
    52: (1.5, {"front": True, "back": True, "left": True, "right": False},
         "damage_status/flood_height from address_assessments.json pipeline (DS2, ~1.5ft "
         "exterior water depth, medium confidence). After-set includes 2023-10-01 "
         "street-level front (N) and left photos (matching the before-photo angles) "
         "plus a 2023-11-01 back photo, all showing the brick/sandstone building intact "
         "with no visible structural, cladding, or fenestration damage compared to the "
         "before photos; right facade only covered by a 2023-07-16 aerial during-flood "
         "shot showing floodwater in the adjacent street but no facade-level detail for "
         "this building. Roof intact in all views. Foundation/basement not visually "
         "inspected."),
    48: (2.0, {"front": True, "back": False, "left": False, "right": False},
         "damage_status/flood_height from address_assessments.json pipeline (DS2, ~2.0ft "
         "exterior water depth, medium confidence). A 2024-06-12 street-level front (N) "
         "photo (same angle as the before-photo) shows the brick Queen Anne house intact "
         "with no visible structural, cladding, or fenestration damage; back/left/right "
         "not covered by after-photos. The 2023-07-11 after-photo is a wide-area aerial "
         "in which this building cannot be individually distinguished for facade-level "
         "assessment. Roof intact. Foundation/basement not visually inspected."),
    28: (3.0, {"front": False, "back": False, "left": False, "right": False},
         "damage_status/flood_height from address_assessments.json pipeline (DS2, ~3.0ft "
         "exterior water depth, medium confidence). All three after-photos (2023-07-16) "
         "are wide-area aerials taken during active flooding; the granite/marble "
         "building appears structurally intact with no visible collapse and the roof "
         "unchanged from the before photos, but no street-level after-photos are "
         "available for any facade - wall/cladding/fenestration columns marked 'un' for "
         "all sides. Floodwater is visible in the surrounding area but the water line at "
         "this building's exterior walls is not determinable from the available aerial "
         "angles. Foundation/basement not visually inspected."),
}


def v(visible):
    return 0 if visible else "un"


for row, (flood_ft, vis, notes) in ROWS.items():
    front_orient = ws.cell(row=row, column=col["front_elevation_orientation"] + 1).value
    cv = cardinal_visibility(front_orient, vis)

    values = {
        "flood_height_building": flood_ft,
        "hazards_present_u": "flood, rain",
        "status_u": "moderate",
        "rainwater_ingress_damage_rating_u": "un",
        "wind_damage_rating_u": 0,
        "damage_indicator_u": 2,
        "degree_of_damage_u": 2,
        "rain_damage_details_u": "un",
        "wind_damage_details_u": "not_applicable",
        "roof_structure_damage_u": 0,
        "roof_structure_damage_u_per": 0,
        "roof_substrate_damage_u": 0,
        "roof_substrate_damage_per": 0,
        "foundation_failure_u": 0,
        "foundation_failure_per_u": "un",
        "wall_structure_damage_per_front": v(vis["front"]),
        "wall_structure_damage_per_back": v(vis["back"]),
        "wall_structure_damage_per_left": v(vis["left"]),
        "wall_structure_damage_per_right": v(vis["right"]),
        "wall_structure_damage_u": 0,
        "wall_structure_damage_n": v(cv["n"]),
        "wall_structure_damage_s": v(cv["s"]),
        "wall_structure_damage_e": v(cv["e"]),
        "wall_structure_damage_w": v(cv["w"]),
        "wall_substrate_damage_per_front": "un",
        "wall_substrate_damage_per_back": "un",
        "wall_substrate_damage_per_right": "un",
        "wall_substrate_damage_per_left": "un",
        "wall_substrate_damage_n": "un",
        "wall_substrate_damage_s": "un",
        "wall_substrate_damage_e": "un",
        "wall_substrate_damage_w": "un",
        "wall_substrate_damage_u": "un",
        "wall_cladding_damage_per_front": v(vis["front"]),
        "wall_cladding_damage_per_back": v(vis["back"]),
        "wall_cladding_damage_per_right": v(vis["right"]),
        "wall_cladding_damage_per_left": v(vis["left"]),
        "wall_cladding_damage_n": v(cv["n"]),
        "wall_cladding_damage_s": v(cv["s"]),
        "wall_cladding_damage_e": v(cv["e"]),
        "wall_cladding_damage_w": v(cv["w"]),
        "damaged_fenesteration_per_front": v(vis["front"]),
        "damaged_fenesteration_per_back": v(vis["back"]),
        "damaged_fenesteration_per_right": v(vis["right"]),
        "damaged_fenesteration_per_left": v(vis["left"]),
        "wall_fenestration_damage_per_n": v(cv["n"]),
        "wall_fenestration_damage_per_s": v(cv["s"]),
        "wall_fenestration_damage_per_e": v(cv["e"]),
        "wall_fenestration_damage_per_w": v(cv["w"]),
        "soffit_damage_per_u": "not_applicable",
        "fascia_damage_per_u": 0,
        "piles_damage_u": "not_applicable",
        "foundation_damage_cause_u": "ot",
        "foundation_damage_u": 1,
        "foundation_damage_u_per": "un",
        "damage_status": 2,
        "Notes": notes,
    }

    for name, val in values.items():
        ws.cell(row=row, column=col[name] + 1).value = val

wb.save(XLSX)
print("done")
