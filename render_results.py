#!/usr/bin/env python3
"""
render_results.py — Read DamageAssessments sheet and write an interactive results.html
with toggleable row coloring, column picker, and always-visible scrollbar.
"""
from __future__ import annotations
from pathlib import Path
from urllib.parse import quote
import openpyxl

BASE_DIR   = Path(__file__).parent
EXCEL_PATH = BASE_DIR / "Montpelier_Flood_DataInput.xlsx"
IMAGE_DIR  = BASE_DIR / "Image gallery"
OUT_PATH   = BASE_DIR / "results.html"

DS_LABELS = {0: "DS 0 — None", 1: "DS 1 — Minor", 2: "DS 2 — Moderate",
             3: "DS 3 — Extensive", 4: "DS 4 — Complete"}
DS_COLOR   = {0: "#d4edda", 1: "#fff3cd", 2: "#fde8c8", 3: "#f8d7da", 4: "#f5c6cb", None: "#f0f0f0"}
CONF_COLOR = {"high": "#d4edda", "medium": "#fff3cd", "low": "#f8d7da"}

# (index, label, visible_by_default)
COLUMNS = [
    (0,  "Image",              True),
    (1,  "File",               False),
    (2,  "Building",           False),
    (3,  "Address (Claude)",   True),
    (4,  "Damage State",       True),
    (5,  "Confidence",         False),
    (6,  "Water Depth",        True),
    (7,  "Image Type",         False),
    (8,  "Assessable",         True),
    (9,  "Rationale",          True),
    (10, "Caveats",            True),
    (11, "Matched Address",    False),
    (12, "Model",              False),
]


def thumb_src(filename: str) -> str:
    if not (IMAGE_DIR / filename).exists():
        return ""
    return f"Image%20gallery/{quote(filename)}"


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    ws = wb["DamageAssessments"]

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers)}
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(r)]

    tbody_rows = []
    for r in rows:
        img_file   = r[col["image_file"]] or ""
        bldg       = r[col["identified_building"]] or "—"
        addr       = r[col["identified_address"]] or "—"
        ds         = r[col["damage_state"]]
        conf       = (r[col["confidence"]] or "").lower()
        depth      = r[col["water_depth_ft_exterior"]]
        img_type   = r[col["image_type"]] or "—"
        assessable = r[col["is_assessable"]]
        rationale  = r[col["rationale"]] or ""
        caveats    = r[col["caveats"]] or ""
        matched    = r[col["matched_address"]] or "—"
        model_val  = r[col["model"]] or ""

        ds_label  = DS_LABELS.get(ds, "Not assessable") if ds is not None else "Not assessable"
        depth_str = f"{depth:.1f} ft" if depth is not None else "—"
        ds_color  = DS_COLOR.get(ds, DS_COLOR[None])
        cf_color  = CONF_COLOR.get(conf, "#f0f0f0")
        src       = thumb_src(img_file)
        thumb     = f'<img src="{src}" class="thumb" alt="">' if src else ""

        cells = [
            thumb,
            f'<span class="fname">{img_file}</span>',
            bldg,
            addr,
            f'<span class="pill" style="background:{ds_color}">{ds_label}</span>',
            f'<span class="pill" style="background:{cf_color}">{conf}</span>',
            depth_str,
            img_type,
            "Yes" if assessable else "No",
            f'<span class="long">{rationale}</span>',
            f'<span class="long">{caveats}</span>',
            matched,
            f'<span class="model">{model_val}</span>',
        ]

        tds = "\n".join(f'  <td data-col="{i}">{v}</td>' for i, v in enumerate(cells))
        tbody_rows.append(
            f'<tr class="data-row" data-ds-color="{ds_color}" data-cf-color="{cf_color}">\n{tds}\n</tr>'
        )

    n_total      = len(rows)
    n_assessable = sum(1 for r in rows if r[col["is_assessable"]] and r[col["damage_state"]] is not None)
    n_matched    = sum(1 for r in rows if r[col["matched_address"]])
    n_flagged    = sum(1 for r in rows if (r[col["confidence"]] or "").lower() == "low"
                       or not r[col["is_assessable"]])

    col_checkboxes = "\n".join(
        f'<label class="col-check"><input type="checkbox" data-col="{i}" '
        f'{"checked" if vis else ""} onchange="toggleCol({i},this.checked)"> {lbl}</label>'
        for i, lbl, vis in COLUMNS
    )

    def th_style(vis):
        return "" if vis else ' style="display:none"'

    thead_ths = "\n".join(
        f'    <th data-col="{i}"{th_style(vis)}>{lbl}</th>'
        for i, lbl, vis in COLUMNS
    )

    # pre-hide columns that are off by default
    hidden_init = "\n".join(
        f"  toggleCol({i}, false);"
        for i, _, vis in COLUMNS if not vis
    )

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Flood Damage Assessment Results</title>
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
        background: #f5f6fa; color: #222; padding: 20px; font-size: 13px; }}
h1 {{ font-size: 1.4rem; margin-bottom: 4px; }}
.subtitle {{ color: #666; font-size: 0.85rem; margin-bottom: 18px; }}

.summary {{ display: flex; gap: 12px; margin-bottom: 18px; flex-wrap: wrap; }}
.stat {{ background: #fff; border-radius: 8px; padding: 10px 18px;
         box-shadow: 0 1px 4px rgba(0,0,0,.08); text-align: center; min-width: 100px; }}
.stat .n {{ font-size: 1.7rem; font-weight: 700; color: #1a73e8; }}
.stat .label {{ font-size: 0.7rem; color: #666; }}

.toolbar {{ display: flex; align-items: flex-start; gap: 20px; margin-bottom: 12px; flex-wrap: wrap; }}

.color-section label.section-label,
.col-section label.section-label {{ font-size: 0.78rem; font-weight: 700; color: #444;
                                     display: block; margin-bottom: 6px; }}
.btn-group {{ display: flex; gap: 6px; flex-wrap: wrap; }}
.toggle-btn {{ padding: 4px 13px; border: 2px solid #ddd; border-radius: 20px;
               background: #fff; cursor: pointer; font-size: 0.75rem; transition: all .15s; }}
.toggle-btn.active {{ border-color: #1a73e8; background: #1a73e8; color: #fff; }}
.toggle-btn:hover:not(.active) {{ border-color: #888; }}

.col-picker {{ display: flex; flex-wrap: wrap; gap: 6px 14px; max-width: 600px; }}
.col-check {{ display: flex; align-items: center; gap: 4px; font-size: 0.75rem;
              cursor: pointer; user-select: none; }}
.col-check input {{ cursor: pointer; }}

.filter-bar {{ margin-bottom: 10px; }}
.filter-bar input {{ padding: 5px 11px; border: 1px solid #ddd; border-radius: 6px;
                     font-size: 0.78rem; width: 240px; }}

.legend {{ display: flex; gap: 10px; flex-wrap: wrap; margin-bottom: 10px; font-size: 0.72rem; }}
.legend-item {{ display: flex; align-items: center; gap: 4px; }}
.swatch {{ width: 11px; height: 11px; border-radius: 3px; flex-shrink: 0; }}

/* Always-visible scrollbar */
.table-wrap {{ overflow-x: scroll; border-radius: 10px;
               box-shadow: 0 1px 6px rgba(0,0,0,.1); }}
table {{ border-collapse: collapse; background: #fff; font-size: 0.78rem; white-space: nowrap; }}
thead th {{ background: #1f4e79; color: #fff; padding: 8px 10px; text-align: left;
            font-size: 0.72rem; font-weight: 600; position: sticky; top: 0; z-index: 2; }}
tbody tr {{ border-bottom: 1px solid #eee; transition: background .1s; }}
tbody tr:hover {{ filter: brightness(0.95); }}
td {{ padding: 7px 10px; vertical-align: top; }}
.thumb {{ width: 68px; height: 50px; object-fit: cover; border-radius: 4px; display: block; }}
.fname {{ display: inline-block; max-width: 140px; overflow: hidden;
          text-overflow: ellipsis; white-space: nowrap; color: #888; font-size: 0.7rem;
          vertical-align: middle; }}
.long {{ display: inline-block; max-width: 260px; white-space: normal;
         line-height: 1.4; font-size: 0.76rem; }}
.model {{ color: #aaa; font-size: 0.7rem; }}
.pill {{ display: inline-block; padding: 2px 7px; border-radius: 4px;
         font-size: 0.72rem; font-weight: 600; }}
</style>
</head>
<body>

<h1>Montpelier Flood — Damage Assessment Baseline</h1>
<p class="subtitle">Model: claude-sonnet-4-6 &nbsp;·&nbsp; {n_total} images processed</p>

<div class="summary">
  <div class="stat"><div class="n">{n_total}</div><div class="label">Total images</div></div>
  <div class="stat"><div class="n">{n_assessable}</div><div class="label">Assessable</div></div>
  <div class="stat"><div class="n">{n_matched}</div><div class="label">Excel matches</div></div>
  <div class="stat"><div class="n">{n_flagged}</div><div class="label">Flagged</div></div>
</div>

<div class="toolbar">
  <div class="color-section">
    <label class="section-label">Color rows by</label>
    <div class="btn-group">
      <button class="toggle-btn active" onclick="setMode('none',this)">None</button>
      <button class="toggle-btn" onclick="setMode('ds',this)">Damage State</button>
      <button class="toggle-btn" onclick="setMode('cf',this)">Confidence</button>
    </div>
  </div>
  <div class="col-section">
    <label class="section-label">Columns</label>
    <div class="col-picker">
{col_checkboxes}
    </div>
  </div>
</div>

<div class="legend" id="legend" style="display:none"></div>

<div class="filter-bar">
  <input type="text" id="filter" placeholder="Filter by address, building, damage state…" oninput="filterRows()">
</div>

<div class="table-wrap">
<table id="results-table">
<thead><tr>
{thead_ths}
</tr></thead>
<tbody>
{"".join(tbody_rows)}
</tbody>
</table>
</div>

<script>
const DS_LEGEND = [
  ["#d4edda","DS 0 — None"],["#fff3cd","DS 1 — Minor"],["#fde8c8","DS 2 — Moderate"],
  ["#f8d7da","DS 3 — Extensive"],["#f5c6cb","DS 4 — Complete"],["#f0f0f0","Not assessable"]
];
const CF_LEGEND = [["#d4edda","High"],["#fff3cd","Medium"],["#f8d7da","Low"]];

let currentMode = 'none';

function setMode(mode, btn) {{
  currentMode = mode;
  document.querySelectorAll('.toggle-btn').forEach(b => b.classList.remove('active'));
  btn.classList.add('active');
  document.querySelectorAll('.data-row').forEach(row => {{
    row.style.background = mode === 'ds' ? row.dataset.dsColor
                         : mode === 'cf' ? row.dataset.cfColor : '';
  }});
  const el = document.getElementById('legend');
  if (mode === 'none') {{ el.style.display = 'none'; return; }}
  const items = mode === 'ds' ? DS_LEGEND : CF_LEGEND;
  el.innerHTML = items.map(([c,l]) =>
    `<div class="legend-item"><div class="swatch" style="background:${{c}}"></div>${{l}}</div>`
  ).join('');
  el.style.display = 'flex';
}}

function toggleCol(idx, visible) {{
  document.querySelectorAll(`[data-col="${{idx}}"]`).forEach(el => {{
    el.style.display = visible ? '' : 'none';
  }});
}}

function filterRows() {{
  const q = document.getElementById('filter').value.toLowerCase();
  document.querySelectorAll('.data-row').forEach(row => {{
    row.style.display = row.textContent.toLowerCase().includes(q) ? '' : 'none';
  }});
}}

// Apply default hidden columns on load
{hidden_init}
</script>
</body>
</html>"""

    OUT_PATH.write_text(html, encoding="utf-8")
    print(f"Written: {OUT_PATH}  ({n_total} rows)")


if __name__ == "__main__":
    main()
