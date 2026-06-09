#!/usr/bin/env python3
from __future__ import annotations
"""
assess_damage.py — Flood damage image assessment pipeline

Reads images from 'Image gallery/', queries the Claude Vision API for structured
damage state assessments (0–4 scale), then writes results into two places:
  1. A new 'DamageAssessments' sheet in Montpelier_Flood_DataInput.xlsx (always)
  2. The 'damage_status' column in BuidlingAttributes (when address matches a row)

Usage:
    ANTHROPIC_API_KEY=sk-ant-... python assess_damage.py
    ANTHROPIC_API_KEY=sk-ant-... python assess_damage.py --model claude-opus-4-7
    ANTHROPIC_API_KEY=sk-ant-... python assess_damage.py --image "Fire and Ambulance Department.jpg"

Dependencies:
    pip install anthropic openpyxl
    (AVIF files are converted automatically on macOS via sips — no extra install needed)
"""

import argparse
import base64
import json
import os
import subprocess
import sys
import tempfile
from difflib import SequenceMatcher
from pathlib import Path

import anthropic
from anthropic import AnthropicVertex
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
IMAGE_DIR = BASE_DIR / "Image gallery"
EXCEL_PATH = BASE_DIR / "Montpelier_Flood_DataInput.xlsx"

# ── API config ─────────────────────────────────────────────────────────────────
DEFAULT_MODEL = "claude-sonnet-4-6"
VERTEX_PROJECT  = os.environ.get("ANTHROPIC_VERTEX_PROJECT_ID", "up-ems-hdr-dsc")
VERTEX_REGION   = os.environ.get("CLOUD_ML_REGION", "us-east5")

# ── Image format support ───────────────────────────────────────────────────────
MEDIA_TYPES = {
    ".jpg":  "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png":  "image/png",
    ".gif":  "image/gif",
    ".webp": "image/webp",
}

# ── Damage state definitions (passed verbatim to the model) ───────────────────
DAMAGE_DEFINITIONS = """\
Damage state scale (0–4) for flood damage assessment:

0 — None: Insignificant damage below first-floor elevation. Water enters crawlspace/basement,
    touches foundation only. Minor damage to garage interiors. No sewer backup into living area.

1 — Minor: Water touches floor joists up to minor water entry. Damage to carpets, pads,
    baseboards, flooring. External AC unit damage if not elevated. No drywall damage; potential
    minor mold on subfloor. Minor sewer backup possible.

2 — Moderate: Partial drywall damage. Damage to base electrical outlets, water heater, furnace.
    Complete damage to first-floor appliances, furniture, lower cabinets. Doors/windows may need
    replacement. Major sewer backup and mold likely.

3 — Extensive: Damage to non-structural components throughout entire building including upper
    stories. Mid-wall electrical destroyed. Upper cabinets destroyed. Studs reusable but some
    damaged. Major sewer backup and mold. Upper-floor contents also damaged.

4 — Complete/Collapsed: Significant structural damage to studs, trusses, joists. All interiors
    destroyed. Roof components damaged. Foundation may have shifted. Building must be demolished."""

SYSTEM_PROMPT = f"""\
You are a structural engineer specializing in flood damage assessment.
Analyze the provided flood image and return a damage assessment using the scale below.
Respond ONLY with a single valid JSON object — no markdown fences, no commentary.

{DAMAGE_DEFINITIONS}

Return exactly these fields:
{{
  "building_name": "best-guess building name, or null",
  "address": "best-guess street address in Montpelier VT, or null if cannot determine",
  "damage_state": integer 0–4, or null if this image cannot support a building-level assessment,
  "confidence": "high" | "medium" | "low",
  "water_depth_ft_exterior": float — estimated water depth in feet at building exterior, or null,
  "image_type": "during_flood" | "post_flood" | "aerial_overview" | "unclear",
  "is_assessable": true if a specific building can be assessed; false for pure aerials or
                   unidentifiable scenes where no individual building is the clear subject,
  "rationale": "1–2 sentence explanation citing specific visual evidence (water line, signage, etc.)",
  "caveats": "key limitations — e.g. exterior only, second floor not visible, oblique angle"
}}"""


# ── Image loading ──────────────────────────────────────────────────────────────

def load_image_as_base64(path: Path) -> tuple[str, str]:
    """Return (base64_data, media_type). Converts AVIF→PNG on macOS via sips."""
    suffix = path.suffix.lower()

    if suffix == ".avif":
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
            tmp_path = Path(tmp.name)
        result = subprocess.run(
            ["sips", "-s", "format", "png", str(path), "--out", str(tmp_path)],
            capture_output=True,
        )
        if result.returncode != 0:
            raise RuntimeError(f"sips conversion failed for {path.name}: {result.stderr.decode()}")
        data = tmp_path.read_bytes()
        tmp_path.unlink()
        return base64.standard_b64encode(data).decode(), "image/png"

    if suffix not in MEDIA_TYPES:
        raise ValueError(f"Unsupported format: {suffix}")

    return base64.standard_b64encode(path.read_bytes()).decode(), MEDIA_TYPES[suffix]


# ── Claude API call ────────────────────────────────────────────────────────────

def assess_image(client: anthropic.Anthropic, image_path: Path, model: str) -> dict:
    """Call Claude Vision API and return the parsed JSON assessment dict."""
    print(f"  Assessing: {image_path.name}")

    b64_data, media_type = load_image_as_base64(image_path)

    message = client.messages.create(
        model=model,
        max_tokens=1024,
        system=SYSTEM_PROMPT,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": media_type,
                            "data": b64_data,
                        },
                    },
                    {
                        "type": "text",
                        "text": (
                            "Assess the flood damage visible in this image. "
                            "Focus on the building(s) most prominent in the frame. "
                            "If you can read any signage, street signs, or address numbers, "
                            "use them to identify the location."
                        ),
                    },
                ],
            }
        ],
    )

    raw = message.content[0].text.strip()

    try:
        result = json.loads(raw)
    except json.JSONDecodeError:
        # Strip accidental markdown fences if the model added them
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        result = json.loads(raw.strip())

    result["image_file"] = image_path.name
    result["model"] = model
    return result


# ── Excel matching ─────────────────────────────────────────────────────────────

def _similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


def find_matching_row(ws_buildings, assessment: dict) -> int | None:
    """
    Try to match assessment address/name to a row in BuidlingAttributes.
    Returns 1-based row index, or None if no confident match found.
    Threshold: 0.75 similarity on address string.
    """
    addr_guess = assessment.get("address") or ""
    name_guess = assessment.get("building_name") or ""

    if not addr_guess and not name_guess:
        return None

    best_row, best_score = None, 0.0

    for row_idx in range(2, ws_buildings.max_row + 1):
        row_addr = ws_buildings.cell(row_idx, 9).value or ""   # complete address
        row_name = ws_buildings.cell(row_idx, 11).value or ""  # building_name_current

        score = max(
            _similarity(addr_guess, str(row_addr)) if addr_guess and row_addr else 0,
            _similarity(name_guess, str(row_name)) if name_guess and row_name else 0,
        )

        if score > best_score:
            best_score = score
            best_row = row_idx

    return best_row if best_score >= 0.75 else None


# ── Output sheet setup ─────────────────────────────────────────────────────────

OUTPUT_COLS = [
    "image_file", "identified_building", "identified_address",
    "damage_state", "confidence", "water_depth_ft_exterior",
    "image_type", "is_assessable", "rationale", "caveats",
    "matched_excel_row", "matched_address", "model",
]

HEADER_FILL   = PatternFill("solid", start_color="1F4E79")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
CELL_FONT     = Font(name="Arial", size=10)
FLAGGED_FILL  = PatternFill("solid", start_color="FFF2CC")  # yellow — needs review


def setup_output_sheet(wb: openpyxl.Workbook) -> openpyxl.worksheet.worksheet.Worksheet:
    if "DamageAssessments" in wb.sheetnames:
        del wb["DamageAssessments"]

    ws = wb.create_sheet("DamageAssessments")

    for col_idx, col_name in enumerate(OUTPUT_COLS, start=1):
        cell = ws.cell(1, col_idx, col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30
    col_widths = [28, 30, 35, 14, 12, 22, 16, 14, 60, 45, 18, 35, 22]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    return ws


def write_assessment_row(ws_out, ws_buildings, row_num: int, assessment: dict) -> None:
    matched_row = find_matching_row(ws_buildings, assessment)
    matched_addr = (
        ws_buildings.cell(matched_row, 9).value if matched_row else None
    )

    values = [
        assessment.get("image_file"),
        assessment.get("building_name"),
        assessment.get("address"),
        assessment.get("damage_state"),
        assessment.get("confidence"),
        assessment.get("water_depth_ft_exterior"),
        assessment.get("image_type"),
        assessment.get("is_assessable"),
        assessment.get("rationale"),
        assessment.get("caveats"),
        matched_row,
        matched_addr,
        assessment.get("model"),
    ]

    needs_flag = (
        assessment.get("confidence") == "low"
        or not assessment.get("is_assessable")
        or assessment.get("damage_state") is None
    )

    for col_idx, val in enumerate(values, start=1):
        cell = ws_out.cell(row_num, col_idx, val)
        cell.font = CELL_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if needs_flag:
            cell.fill = FLAGGED_FILL

    # Write back to BuidlingAttributes if we found a match
    if matched_row and assessment.get("damage_state") is not None:
        ws_buildings.cell(matched_row, 244).value = assessment["damage_state"]  # damage_status
        if assessment.get("water_depth_ft_exterior") is not None:
            ws_buildings.cell(matched_row, 3).value = assessment["water_depth_ft_exterior"]  # flood_height_building


# ── Main ───────────────────────────────────────────────────────────────────────

def collect_images(single: str | None) -> list[Path]:
    if single:
        p = IMAGE_DIR / single
        if not p.exists():
            sys.exit(f"Image not found: {p}")
        return [p]

    supported = set(MEDIA_TYPES) | {".avif"}
    images = sorted(
        p for p in IMAGE_DIR.iterdir()
        if p.is_file() and p.suffix.lower() in supported
    )
    if not images:
        sys.exit(f"No supported images found in {IMAGE_DIR}")
    return images


def main():
    parser = argparse.ArgumentParser(description="Flood damage image assessment pipeline")
    parser.add_argument("--model", default=DEFAULT_MODEL, help="Claude model ID")
    parser.add_argument("--image", default=None, help="Assess a single image by filename")
    args = parser.parse_args()

    images = collect_images(args.single if hasattr(args, "single") else args.image)
    print(f"Images to process: {len(images)}")
    print(f"Model: {args.model}")
    print(f"Vertex project: {VERTEX_PROJECT} / {VERTEX_REGION}")
    print(f"Excel: {EXCEL_PATH.name}\n")

    client = AnthropicVertex(project_id=VERTEX_PROJECT, region=VERTEX_REGION)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws_buildings = wb["BuidlingAttributes"]
    ws_out = setup_output_sheet(wb)

    results = []
    for img_path in images:
        try:
            assessment = assess_image(client, img_path, args.model)
            results.append(assessment)

            state = assessment.get("damage_state")
            conf  = assessment.get("confidence", "?")
            addr  = assessment.get("address") or "address unknown"
            assessable = assessment.get("is_assessable", True)

            if not assessable:
                print(f"    → Not assessable (aerial/unclear scene)")
            else:
                print(f"    → DS {state} ({conf} confidence) | {addr}")

        except Exception as exc:
            print(f"    ERROR: {exc}")
            results.append({
                "image_file": img_path.name,
                "error": str(exc),
                "is_assessable": False,
                "damage_state": None,
                "confidence": "low",
            })

    print(f"\nWriting {len(results)} results to Excel...")
    for row_num, assessment in enumerate(results, start=2):
        write_assessment_row(ws_out, ws_buildings, row_num, assessment)

    wb.save(EXCEL_PATH)

    assessable = [r for r in results if r.get("is_assessable") and r.get("damage_state") is not None]
    matched    = [r for r in assessable if find_matching_row(ws_buildings, r)]
    flagged    = [r for r in results if r.get("confidence") == "low" or not r.get("is_assessable")]

    print(f"\nDone.")
    print(f"  Total images processed : {len(results)}")
    print(f"  Assessable buildings   : {len(assessable)}")
    print(f"  Matched to Excel rows  : {len(matched)}")
    print(f"  Flagged for review     : {len(flagged)}")
    print(f"  Output sheet           : DamageAssessments tab in {EXCEL_PATH.name}")


if __name__ == "__main__":
    main()
